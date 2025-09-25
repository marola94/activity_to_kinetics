import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from . import validation


def load_measurements_file(file_path):
    # Load the workbook and the specific cell
    workbook = load_workbook(file_path, data_only=True)  # Load with values only
    sheet = workbook['Measurements']  # Use the active sheet

    ########################################
    # Dictionary for all measurement and conversion parameters 
    all_params = {
        'Measurement unit': None,
        'Time unit': None,
        'Desired unit': None,
        'Calibration curve slope': None,
        'Calibration curve slope unit': None,
        'Calibration curve intercept': None,
        'Extinction coefficient': None,
        'Extinction coefficient unit': None,
        'Path length': None,
        'Path length unit': None
    }

    # Iterate through the excel sheet to find and define parameters
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == 'Measurement unit:':
                all_params['Measurement unit'] = sheet.cell(row=cell.row, column=cell.column + 1).value

            elif cell.value == 'Time unit:':
                unit_cell = sheet.cell(row=cell.row, column=cell.column + 1).value
                all_params['Time unit'] = np.nan if unit_cell is None or str(unit_cell).strip().lower() == 'Time format' else str(unit_cell).strip()

            elif cell.value == 'Desired unit:':
                unit_cell = sheet.cell(row=cell.row, column=cell.column + 1).value
                all_params['Desired unit'] = np.nan if unit_cell is None or str(unit_cell).strip().lower() == 'no conversion' else str(unit_cell).strip()

            elif cell.value == 'Calibration curve slope:':
                value_cell = sheet.cell(row=cell.row, column=cell.column + 1).value
                unit_cell = sheet.cell(row=cell.row, column=cell.column + 2).value
                all_params['Calibration curve slope'] = np.nan if value_cell is None or str(value_cell).strip() == '' else float(value_cell)
                all_params['Calibration curve slope unit'] = np.nan if unit_cell is None or str(unit_cell).strip().lower() == 'no conversion' else str(unit_cell).strip()

            elif cell.value == 'Calibration curve intercept:':
                value_cell = sheet.cell(row=cell.row, column=cell.column + 1).value
                all_params['Calibration curve intercept'] = np.nan if value_cell is None or str(value_cell).strip() == '' else float(value_cell)

            elif cell.value == 'Extinction coefficient:':
                value_cell = sheet.cell(row=cell.row, column=cell.column + 1).value
                unit_cell = sheet.cell(row=cell.row, column=cell.column + 2).value
                all_params['Extinction coefficient'] = np.nan if value_cell is None or str(value_cell).strip() == '' else float(value_cell)
                all_params['Extinction coefficient unit'] = np.nan if unit_cell is None or str(unit_cell).strip().lower() == 'no conversion' else str(unit_cell).strip()

            elif cell.value == 'Path length:':
                value_cell = sheet.cell(row=cell.row, column=cell.column + 1).value
                unit_cell = sheet.cell(row=cell.row, column=cell.column + 2).value
                all_params['Path length'] = np.nan if value_cell is None or str(value_cell).strip() == '' else float(value_cell)
                all_params['Path length unit'] = np.nan if unit_cell is None or str(unit_cell).strip().lower() == 'no conversion' else str(unit_cell).strip()

        # Stop loopet hvis alle værdier er fundet
        if all(v is not None for v in all_params.values()):
            break

    ########################################

    ###########################################################
    # Define reading frame of measurement 
    # Find the location in excel sheet with 'Title' in the cell, which indicates where the read should start from
    start_row = None
    for row in sheet.iter_rows():
        for cell in row:
            if str(cell.value).lower() == 'experiment title':
                start_row = cell.row # defined start row
                start_column = cell.column # defined start column
                break

        if start_row and start_column:
            break

    # Find the last used row and column in the excel sheet
    max_row = sheet.max_row
    max_col = sheet.max_column

    for r in range(start_row + 1, max_row + 1):
        for c in range(start_column, max_col + 1):
            if sheet.cell(r, c).value is not None:
                end_row = max(start_row, r)
                end_col = max(start_column, c)
    
    # Convert dataframe column index, i.e. numbers, to excel column index, i.e. letters
    start_col_letter = get_column_letter(start_column)
    end_col_letter = get_column_letter(end_col)
    ###########################################################

    ######################################################
    # Read dataframe with only well koordinates and titles
    titles_start = get_column_letter(start_column+1) # Start column to read excel sheet to dataframe 
    titles_df = pd.read_excel(
        file_path,
        sheet_name='Measurements',
        usecols=f"{titles_start}:{end_col_letter}",
        skiprows=start_row - 1,            
        nrows=2,                           
        header=None
    )

    # Modify dataframe and filter columns without defined titles 
    titles_df = pd.DataFrame([titles_df.iloc[0].values], columns=titles_df.iloc[1].values, index=['Title'])
    valid_columns = titles_df.loc['Title'][(titles_df.loc['Title'].notna()) & (titles_df.loc['Title'].astype(str).str.strip() != '')].index
    titles_df = titles_df[valid_columns]
    ######################################################


    ######################################################
    # Read dataframe with measurements 
    data_df = pd.read_excel(
        file_path, 
        sheet_name = 'Measurements',
        usecols=f"{start_col_letter}:{end_col_letter}", 
        skiprows=start_row, 
        nrows=end_row - start_row+1
    )
    
    data_df = data_df.set_index('Time')

    ######################################################
    # Check which parameters are present for converting measurements to units
    conversion_dictionary = validation.validate_parameters(all_params)
    ######################################################

    ######################################################
    # Checking format of time values and converting to array in pint units
    validation.checking_time_format(data_df, conversion_dictionary)
    ######################################################

    ######################################################
    # Remove columns where all values are NaN
    data_df = data_df.dropna(axis=1, how='all')
    ######################################################

    return conversion_dictionary, data_df, titles_df

def load_manual_rate_fitting(file_path:str) -> pd.DataFrame:
    """
    Load settings for fitting rates with manual input from the 'Rate_settings' sheet in the uploaded Excel file (e.g. 'Template_data_file.xlsx')
    """

    workbook = load_workbook(file_path, data_only=True)  # Load with values only
    sheet = workbook['Rate settings']  # Use the active sheet

    ###########################################################
    # Define reading frame 
    # Find the location in excel sheet with 'Well' in the cell, which indicates where the read should start from
    start_row = None
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == 'Well':
                start_row = cell.row # defined start row
                start_column = cell.column # defined start column
                break

        if start_row and start_column:
            break

    # Find the last used row and column in the excel sheet
    end_row = sheet.max_row
    end_col = sheet.max_column
    
    # Convert dataframe column index, i.e. numbers, to excel column index, i.e. letters
    start_col_letter = get_column_letter(start_column)
    end_col_letter = get_column_letter(end_col)
    ###########################################################

    ###########################################################
    # Load datapoint intervals for manual slope fitting
    slope_fitting_df = pd.read_excel(
        file_path,
        sheet_name = 'Rate settings',
        usecols=f"{start_col_letter}:{end_col_letter}", 
        nrows=end_row - start_row,
        skiprows=start_row-1
        )
    
    slope_fitting_df = slope_fitting_df.set_index('Well') # Set well as index
    slope_fitting_df = slope_fitting_df.T # Transpose
    slope_fitting_df = slope_fitting_df.apply(pd.to_numeric, errors='coerce').dropna(axis=1, how='all') # Drop empty columns i.e. no specified datapoint interval
    slope_fitting_df= slope_fitting_df.round(0).astype('Int64') # Convert to integers
    ###########################################################

    return slope_fitting_df

def load_kinetics_settings(file_path:str) -> dict:

    kinetics_settings_df, substrate_conc_unit, enzyme_conc_unit = kinetics_user_inputs(file_path)

    kinetics_settings_dict = create_kinetics_dict(kinetics_settings_df, substrate_conc_unit, enzyme_conc_unit)

    if not kinetics_settings_dict: # Empty due empty settings file 
        return None

    cleaned_models = validation.validating_kinetics_dictionary(kinetics_settings_dict)
    if not cleaned_models: # Empty due to all models were invalid and discarded
        return None 

    return cleaned_models
    

def kinetics_user_inputs(file_path: str):
    workbook = load_workbook(file_path, data_only=True)  # Load with values only
    sheet = workbook['Kinetics settings']  # Use the named sheet

    ###########################################################
    # Define reading frame of measurement 
    # 1) find first 'Model No.'
    start_row = start_col = None
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.strip().lower() == "model no.":
                start_row, start_col = cell.row, cell.column
                break
        if start_row is not None:
            break
    if start_row is None:
        raise ValueError("Could not find 'Model No.' in the sheet.")

    max_row, max_col = sheet.max_row, sheet.max_column

    # 2a) find last non-empty column to the right of start_col
    end_col = start_col
    for c in range(start_col, max_col + 1):
        any_non_empty = False
        for r in range(start_row, max_row + 1):
            v = sheet.cell(r, c).value
            # treat empty strings/whitespace as empty
            if v is not None and (not isinstance(v, str) or v.strip() != ""):
                any_non_empty = True
                break
        if any_non_empty:
            end_col = c

    # 2b) find last non-empty row within [start_col .. end_col]
    end_row = start_row
    for r in range(start_row, max_row + 1):
        any_non_empty = False
        for c in range(start_col, end_col + 1):
            v = sheet.cell(r, c).value
            if v is not None and (not isinstance(v, str) or v.strip() != ""):
                any_non_empty = True
                break
        if any_non_empty:
            end_row = r
    
    # Convert dataframe column index, i.e. numbers, to excel column index, i.e. letters
    start_col_letter = get_column_letter(start_col)
    end_col_letter = get_column_letter(end_col)
    ###########################################################

    ###########################################################
    # Read units defined for substrate concentraion and enzyme concentration
    enzyme_conc_unit = None
    substrate_conc_unit = None

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            v = cell.value
            if isinstance(v, str):
                key = v.strip().lower().rstrip(":")
                if "enzyme concentration unit" in key and enzyme_conc_unit is None:
                    enzyme_conc_unit = sheet.cell(row=cell.row, column=cell.column + 1).value
                elif "substrate concentration unit" in key and substrate_conc_unit is None:
                    substrate_conc_unit = sheet.cell(row=cell.row, column=cell.column + 1).value
    ###########################################################

    ###########################################################
    # Validate units for enzyme and substrate concentration 

    enzyme_conc_unit = validation.validate_unit(enzyme_conc_unit, "enzyme")
    substrate_conc_unit = validation.validate_unit(substrate_conc_unit, "substrate")

    ###########################################################

    ###########################################################
    # Load the reading frame into a DataFrame
    kinetics_settings_df = pd.read_excel(
        file_path,
        sheet_name='Kinetics settings',
        usecols=f"{start_col_letter}:{end_col_letter}",
        nrows=(end_row - start_row),   # header is at start_row
        skiprows=start_row - 1         # skip rows before the header
    )
    ###########################################################

    return kinetics_settings_df, substrate_conc_unit, enzyme_conc_unit

def create_kinetics_dict(kinetics_settings_df: pd.DataFrame, enzyme_conc_unit, substrate_conc_unit) -> dict:
    """
    Convert a 'kinetics_settings_df' frame into a nested dict:
        dict ={"Model No.": {
                "Model type": <str>,
                "Title": <str>,                         # empty string if missing
                "Enzyme concentration": <float|orig>,   # NaN if missing; tries float
                "data": <DataFrame>                     # index = substrate concentration, column 'Wells'
                },
            ...
            }   

        Expected layout in kinetics_settings_df (per model, 2 columns):
        [info_col='Model No.' (or 'Model No..k'), value_col='<model index>']
        Row 0 in value_col: Model type
        Row 1 in value_col: Title (optional)
        Row 2 in value_col: Enzyme concentration
        Row 3: header line: left='Substrate concentration', right='Wells'
        Row 4..: data rows: left=sub conc, right=wells
    """
    kinetics_settings_dict = {}
    df = kinetics_settings_df
    if df is None or df.empty:
        return kinetics_settings_dict

    # Iterate through blocks with settings for the models
    model_cols = [c for c in df.columns if "Model No." in str(c)]

    for col in model_cols:
        info_col = df[col]
        value_col_index = df.columns.get_loc(col) + 1
        value_col = df.iloc[:, value_col_index]

        # --- Find Model type ---
        model_type = None
        for i, val in enumerate(info_col):
            if isinstance(val, str) and val.strip().lower() == "model type":
                cell_val = value_col.iloc[i]
                if pd.notna(cell_val):
                    model_type = str(cell_val).strip()
                break

        # --- Find Title ---
        title = None
        for i, val in enumerate(info_col):
            if isinstance(val, str) and val.strip().lower() == "title":
                cell_val = value_col.iloc[i]
                if pd.notna(cell_val):
                    title = str(cell_val).strip()
                break

        # --- Find Enzyme concentration ---
        enzyme_conc = None
        for i, val in enumerate(info_col):
            if isinstance(val, str) and val.strip().lower() == "enzyme concentration":
                cell_val = value_col.iloc[i]
                if pd.notna(cell_val):
                    enzyme_conc = cell_val
                break

        # ---- Find rows with 'Substrate concentration' and 'Wells' in info_col and value_col respectively ----
        subs_row = None
        wells_row = None

        for i, (left, right) in enumerate(zip(info_col, value_col)):
            if isinstance(left, str) and left.strip().lower() == "substrate concentration":
                subs_row = i
            if isinstance(right, str) and right.strip().lower() == "wells":
                wells_row = i
            if subs_row is not None and wells_row is not None:
                break  # stop as soon as both are found

        
        # Slice all rows below the header rows
        sub_conc_input  = info_col.iloc[subs_row+1:]  if subs_row  is not None else pd.Series([], dtype=object)
        wells_input = value_col.iloc[wells_row+1:] if wells_row is not None else pd.Series([], dtype=object)
        
        if (
            enzyme_conc is not None
            
            or sub_conc_input.apply(lambda x: pd.notna(x) and (not isinstance(x, str) or x.strip() != "")).any()
            or wells_input.apply(lambda x: pd.notna(x) and (not isinstance(x, str) or x.strip() != "")).any()
        ):

            input_df = pd.DataFrame({
                "Substrate concentration": sub_conc_input.values,
                "Wells": wells_input.values
                }).set_index("Substrate concentration").rename_axis(f"Substrate concentration [{substrate_conc_unit}]").dropna(how="all")

            kinetics_settings_dict[f'Model {value_col.name}'] = {
                "Model type": model_type,
                "Title": title,
                f"Enzyme concentration [{enzyme_conc_unit}]": enzyme_conc,
                "data": input_df
            }

    return kinetics_settings_dict



def load_rate_data_file(filename, sheetname=None, wells_to_read=None):
    """
    Load an Excel file and optionally extract specific rows based on well names.
    
    Args:
    filename (str): The path to the Excel file.
    sheetname (str, optional): The name of the sheet to read. If not specified, the default sheet is read.
    wells (list, optional): A list of well names to extract from the dataframe. If not specified, the entire dataframe is returned.
    
    Returns:
    pd.DataFrame: The loaded dataframe, possibly filtered by well names.
    """
    # Load the Excel file
    if sheetname:
        df = pd.read_excel(filename, sheetname=sheetname, index_col=0)
    else:
        df = pd.read_excel(filename, index_col=0)

    wells = df.index
    
    # Filter the dataframe by well names if specified
    if wells_to_read:
        well_list = []
        for p in wells_to_read:
            if p.isalpha() == True:
                for well in wells:
                    for i in str(well):
                        if i.isalpha() == True:
                            if i == p:
                                well_list.append(well)

                            #for j in plate_lines:
                            #    if i == j:
                            #        print('i',i,'j',j)
                            #        print('HERE1',well)
                            #        print(p)
                            #        well_list.append(well)
            else:
                for well in wells:
                    if p == well:
                        well_list.append(well)

        df = df.loc[well_list]
    
    return df




