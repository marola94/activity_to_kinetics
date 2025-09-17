import os
import string
import sys
from openpyxl import load_workbook

def user_interface(settings_dict):

    """
    Console user interface for program instructions. The user specifies if the rates of enzyme activitty or 
    enzyme kinetics should be estimated. 

    """
    print('\n***************************************************************************************************************************')
    print('\n Main menu')

    ######################
    # 1. Program options #
    ######################
    options = ['Estimate rates', 'Estimate kinetics', 'Plot data', 'Exit']
    letters = string.ascii_uppercase[:len(options)]  # Generates A, B, C etc.
    first_letter = letters[0]
    last_letter = letters[-1]

    # Show program options
    print('Select program for your data:')
    for letter, option in zip(letters, options):
        print(f"{letter}) {option}")

    # The user choose an option
    choice = input(f'Choose program option ({first_letter}-{last_letter}): ').strip().upper()
    while choice not in letters[:len(options)]:
        print("Invalid choice. Try again.")
        choice = input(f'Choose program option ({first_letter}-{last_letter}): ').strip().upper()
    
    program_choice = options[letters.index(choice)]  # Get the corresponding value

    if program_choice == 'Estimate rates':
        settings_dict = rate_estimation_interface()

    elif program_choice == 'Estimate kinetics':
        settings_dict = kinetics_estimation_interface()

    elif program_choice == 'Plot data':
        settings_dict = plot_data_interface()

    elif program_choice == "Exit":
        print("\nTerminating program.")
        sys.exit(0)

    return settings_dict
    
def rate_estimation_interface():

    print('------------------------------------------------------------------------------------------------------------------------')
    print("\nConfiguration for estimating rates.")

    settings_dict = {}
    step = 1  # Tracks the current step

    while True:
        if step == 1:
            ###########################
            # Step 1: File to analyze #
            ###########################

            print("\nEnter the path to the file with measurements and times (e.g., 'measurements.xlsx'), type 'Back' to return to the main menu.")
            
            file_path = input("Path to data file: ").strip()

            if file_path.lower() == "back":
                print("Returning to the main menu.")
                return None
            
            elif file_path.lower() == "exit":
                print("\nTerminating program.")
                sys.exit(0)

            #elif os.path.isfile(file_path) and file_path.endswith(".xlsx"):
                

            else:
                if '.xlsx' not in file_path:
                    file_path = f'{file_path}'+'.xlsx'
                
                while os.path.isfile(file_path) == False:
                    print("Invalid file path. Please try again.")
                    file_path = input("Path to data file: ").strip()

                if os.path.isfile(file_path) == True:
                    try:
                        # Save the file path in the dictionary 
                        settings_dict["file_path"] = file_path

                        # Load the workbook and the specific cell
                        workbook = load_workbook(file_path, data_only=True)  # Load with values only
                        sheet = workbook['Measurements']  # Use the active sheet

                        # Read the unit of measurements 
                        specific_cell_value = sheet["C21"].value  # Read value of cell C21
                        if specific_cell_value == "Au":
                            step = 2
                        else:
                            step = 3

                        print(f'\nInserted file path: {file_path}')

                    except Exception as e:
                        print(f"Failed to read the file. Error: {e}")

        if step == 2:
            #################################################
            # Step 2: Convert absorbances to concentrations #
            #################################################

            print("\nConvert absorbances to concentrations?")
            options = ['Yes', 'No', 'Back', 'Back to main menu', 'Exit']
            letters = string.ascii_uppercase[:len(options)]  # Generates A, B, C etc.
            first_letter = letters[0]
            last_letter = letters[-1]

            # Show options
            for letter, option in zip(letters, options):
                print(f"{letter}) {option}")

            # The user chooses an option
            choice = input(f'Select an option ({first_letter}-{last_letter}): ').strip().upper()
            while choice not in letters[:len(options)]:
                print("Invalid choice. Try again.")
                choice = input(f'Select an option ({first_letter}-{last_letter}): ').strip().upper()
            
            if choice == "A":
                settings_dict['convert_absorbances'] = convert_absorbance()
                if settings_dict['convert_absorbances'] == None:
                    step = 2 # Remain at this step 
                elif settings_dict['convert_absorbances']:
                    step = 3  # Move to the next step
            elif choice == "B":
                settings_dict["convert_absorbances"] = False
                step = 3  # Move to the next step
            elif choice == "C":
                step = 1  # Go back to file import
            elif choice == "D":
                print("\nReturning to the main menu.")
                return None
            elif choice == "E":
                print("\nTerminating program.")
                sys.exit(0)
            else:
                print("Invalid choice. Please try again.")

        if step == 3:
            ########################################
            # Step 3: Rates estimation output file #
            ########################################
            
            print("\nThe default file name of the output file for the calculated rates is 'rates_output.xlsx'.")
            print("\nDo you want to define the file name of the output excel file?",
                  " Already existing files will be overwritten,",
                  " and if a file name is not provided then the default file name of the output file will be used.")
            options = ['Yes', 'No', 'Back', 'Back to main menu', 'Exit']
            letters = string.ascii_uppercase[:len(options)]  # Generates A, B, C etc.
            first_letter = letters[0]
            last_letter = letters[-1]

            # Show options
            for letter, option in zip(letters, options):
                print(f"{letter}) {option}")

            # The user chooses an option
            choice = input(f'Select an option ({first_letter}-{last_letter}): ').strip().upper()
            while choice not in letters[:len(options)]:
                print("Invalid choice. Try again.")
                choice = input(f'Select an option ({first_letter}-{last_letter}): ').strip().upper()
            
            if choice == "A":
                
                print("\nOutput filename.")
                output_filename = input(f"\nWrite name of output file for the rates: ").strip()
                invalid_characters = ['<', '>', ':', '"', '/', '|', '?', '*']
                while (output_filename in invalid_characters == True) or (len(output_filename)>218) or (output_filename):
                    print('Invalid input. Please enter a name without the characters <, >, :, ", /, |, ?, * and less than 218 characters',
                          'or type "Back" to go back or "Exit" to quit.')
                    output_filename = input(f"\nWrite name of output file for the rates: ").strip()

                if output_filename.lower() == 'back':
                    continue

                elif output_filename.lower() == 'exit':
                    print("\nTerminating program.")
                    sys.exit(0)

                settings_dict["Output filename"] = output_filename
        
            elif choice == "B":
                settings_dict["Output filename"] = False
                step = 4  # Move to the next step
            
            elif choice == "C":
                step = 1  # Go back to file selection

            elif choice == "D":
                print("\nReturning to the main menu.")
                return None
            elif choice == "E":
                print("\nTerminating program.")
                sys.exit(0)
            else:
                print("Invalid choice. Please try again.")

        if step == 4:
            ####################################
            # Step 4: Manual fitting of slopes #
            ####################################

            print("\nManual rate fitting\n")
            print("\nTo set rates manually, follow the instructions in the sheet 'Setting rates manually' of the input file template.",
                  "\nDo you want to apply the settings for the rates?")
            
            options = ['Yes', 'No', 'Back', 'Back to main menu', 'Exit']
            letters = string.ascii_uppercase[:len(options)]  # Generates A, B, C etc.
            first_letter = letters[0]
            last_letter = letters[-1]

            # Show options
            for letter, option in zip(letters, options):
                print(f"{letter}) {option}")

            # The user chooses an option
            choice = input(f'Select an option ({first_letter}-{last_letter}): ').strip().upper()
            while choice not in letters[:len(options)]:
                print("Invalid choice. Try again.")
                choice = input(f'Select an option ({first_letter}-{last_letter}): ').strip().upper()

            if choice == 'A':
                settings_dict['Manual rate fitting'] = True
                return settings_dict
            
            elif choice == 'B': 
                settings_dict['Manual rate fitting'] = False
                return settings_dict

            elif choice == 'C':
                step = 3

            elif choice == 'D':
                print("\nReturning to the main menu.")
                return None

            elif choice == 'E':
                print("\nTerminating program.")
                sys.exit(0)

        #return settings_dict

def convert_absorbance():

    #########################
    # Absorbance conversion #
    #########################

    abs_conversion_dict = {}

    print("\nWhich unit do you want to convert absorbances to?") 
    options = ['mol/L', 'mmol/L', 'µmol/L', 'nmol/L', 'Back', 'Exit']
    letters = string.ascii_uppercase[:len(options)]  # Generates A, B, C etc.
    first_letter = letters[0]
    last_letter = letters[-1]

    # Show options
    for letter, option in zip(letters, options):
        print(f"{letter}) {option}")

    # The user chooses an option
    choice = input(f'Select an option ({first_letter}-{last_letter}): ').strip().upper()
    while choice not in letters[:len(options)]:
        print("Invalid choice. Try again.")
        choice = input(f'Select an option ({first_letter}-{last_letter}): ').strip().upper()

    if choice == 'A' or 'B' or 'C' or 'D':
        unit_choice = options[letters.index(choice)]  # Get the corresponding value

        print("\nHow do you want to convert absorbances to concentrations?") 
        options = ['Standard curve', 'Lambert-Beers law', 'Back', 'Exit']
        letters = string.ascii_uppercase[:len(options)]  # Generates A, B, C etc.
        first_letter = letters[0]
        last_letter = letters[-1]

        # Show options
        for letter, option in zip(letters, options):
            print(f"{letter}) {option}")

        # The user chooses an option
        choice = input(f'Select an option ({first_letter}-{last_letter}): ').strip().upper()
        while choice not in letters[:len(options)]:
            print("Invalid choice. Try again.")
            choice = input(f'Select an option ({first_letter}-{last_letter}): ').strip().upper()
                
        conversion_choice = options[letters.index(choice)]  # Get the corresponding value

        # Getting parameters for standard curve
        if conversion_choice == 'Standard curve':
            slope_unit_0 = unit_choice.split('/')[0]
            slope_unit_1 = unit_choice.split('/')[1]
            slope_unit = f'{slope_unit_1}/{slope_unit_0}'

            print("\nSlope and intercept of standard curve.")
            slope_input = input(f"\nProvide the slope from linear regression in {slope_unit}: ").strip()
            while (float(slope_input) == False) or (slope_input.lower() != 'back') or (slope_input.lower() != 'exit'):
                print("Invalid input. Please enter a valid number, or type 'Back' to go back or 'Exit' to quit.")
                slope_input = input(f"\nProvide the slope from linear regression in the unit of {slope_unit}: ").strip()

            if slope_input.lower() == 'back':
                return None

            elif slope_input.lower() == 'exit':
                print("\nTerminating program.")
                sys.exit(0)

            intercept_input = input(f"\nProvide the intercept from linear regression: ").strip()
            print(type(intercept_input))
            while (float(intercept_input) == False) or (intercept_input.lower() != 'back') or (intercept_input.lower() != 'exit'):
                print("Invalid input. Please enter a valid number, or type 'Back' to go back or 'Exit' to quit.")
                intercept_input = input(f"\nProvide the intercept from linear regression: ").strip()
                    
            if intercept_input.lower() == 'back':
                return None

            elif intercept_input.lower() == 'exit':
                print("\nTerminating program.")
                sys.exit(0)

            abs_conversion_dict['Unit'] = unit_choice
            abs_conversion_dict['Conversion type'] = conversion_choice
            abs_conversion_dict['Slope'] = slope_input
            abs_conversion_dict['Intercept'] = intercept_input

            return abs_conversion_dict

        # Getting parameters for Lamber-Beers equation
        elif conversion_choice == 'Lambert-Beers law':

            print("\nParameters for Lambert-Beers law.")
            ext_coefficient = input("\nProvide the extincion coefficient of absorbing chemical in M$^{-1}$cm${-1}$: ").strip()
            while (float(ext_coefficient) == False) or (ext_coefficient.lower() != 'back') or (ext_coefficient.lower() != 'exit'):
                print("Invalid input. Please enter a valid number, or type 'Back' to go back or 'Exit' to quit.")
                ext_coefficient = input("\nProvide the extincion coefficient of absorbing chemical in M$^{-1}$cm${-1}$: ").strip()
                    
            if ext_coefficient.lower() == 'back':
                return None

            elif ext_coefficient.lower() == 'exit':
                print("\nTerminating program.")
                sys.exit(0)

            path_length =  input("\nProvide the path length of the instrument in cm").strip()
            while (float(path_length) == False) or (path_length.lower() != 'back') or (path_length.lower() != 'exit'):
                print("Invalid input. Please enter a valid number, or type 'Back' to go back or 'Exit' to quit.")
                path_length =  input("\nProvide the path length of the instrument in cm").strip()
                    
            if path_length.lower() == 'back':
                return None

            elif path_length.lower() == 'exit':
                print("\nTerminating program.")
                sys.exit(0)

            abs_conversion_dict['Unit'] = unit_choice
            abs_conversion_dict['Conversion type'] = conversion_choice
            abs_conversion_dict['Extincion coefficient'] = ext_coefficient
            abs_conversion_dict['Path length'] = path_length

            return abs_conversion_dict
                
        elif conversion_choice.lower() == 'back':
            return None

        elif conversion_choice.lower() == 'exit':
            print("\nTerminating program.")
            sys.exit(0)
        else:
            print("Invalid choice. Please try again.")    

    elif choice == "E":
        step = 2  # Go back to asking if absorbances should be converted

    elif choice == "F":
        print("\nTerminating program.")
        sys.exit(0)

    else:
        print("Invalid choice. Please try again.")

def kinetics_estimation_interface():

    print('------------------------------------------------------------------------------------------------------------------------')
    print("\nConfiguration for estimating kinetics.")

    settings_dict = {}
    step = 1  # Tracks the current step

    while True:
        if step == 1:
            ######################################
            # Step 1: File with rates to analyze #
            ######################################
            print("\nEnter the path to the file with rates (e.g., 'rates_output.xlsx'), type 'Back' to return to the main menu.")
            
            file_path = input("Path to data file: ").strip()

            if file_path.lower() == "back":
                print("Returning to the main menu.")
                return None
            
            elif file_path.lower() == "exit":
                print("\nTerminating program.")
                sys.exit(0)

            elif os.path.isfile(file_path) and file_path.endswith(".xlsx"):
                try:
                    # Save the file path in the dictionary 
                    settings_dict["file_path"] = file_path

                    # Load the workbook and the specific cell
                    workbook = load_workbook(file_path, data_only=True)  # Load with values only
                    sheet = workbook['Measurements']  # Use the active sheet

                    # Read the unit of measurements 
                    specific_cell_value = sheet["C21"].value  # Read value of cell C21
                    if specific_cell_value == "Au":
                        step = 2
                    else:
                        step = 3

                    print('\nInserted file path: {file_path}')

                except Exception as e:
                    print(f"Failed to read the file. Error: {e}")

            else:
                print("Invalid file path. Please try again.")

        #elif step == 2:
            ###############################
            # Step 2: Removal of outliers #
            ###############################

        #elif step == 3:
            ########################################
            # Step 3: Rates estimation output file #
            ########################################

        







    return settings_dict

def plot_data_interface():

    return settings_dict